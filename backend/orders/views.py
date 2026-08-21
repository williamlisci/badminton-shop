from calendar import monthrange
from datetime import date, datetime, time, timedelta
from decimal import Decimal
import csv
from io import StringIO
from io import BytesIO

from django.db.models import Count, DecimalField, F, Q, Sum, OuterRef, Subquery
from django.http import HttpResponse
from django.db.models.functions import Coalesce, TruncDate, TruncMonth
from django.utils import timezone
from rest_framework import generics, status, viewsets
from rest_framework.permissions import AllowAny, IsAuthenticated
from accounts.permissions import OrderManagerPermission, ReportPermission
from accounts.audit import record_audit
from accounts.models import AuditLog
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView

from products.models import Category, Product
from .models import Customer, DiscountCode, Order, OrderItem, OrderStatusHistory
from .serializers import (
    AdminOrderSerializer,
    AdminCustomerSerializer,
    OrderCreateSerializer,
    OrderSerializer,
    DiscountValidationSerializer,
)
from rest_framework import serializers


class DiscountCodeSerializer(serializers.ModelSerializer):
    is_active = serializers.BooleanField(required=False, default=True)
    max_uses = serializers.IntegerField(required=False, allow_null=True, min_value=0)
    used_count = serializers.IntegerField(read_only=True)

    product_ids = serializers.PrimaryKeyRelatedField(
        source='products', many=True, queryset=Product.objects.all(),
        required=False,
    )
    category_ids = serializers.PrimaryKeyRelatedField(
        source='categories', many=True,
        queryset=Category.objects.all(),
        required=False,
    )

    class Meta:
        model = DiscountCode
        fields = [
            'id', 'code', 'discount_type', 'value', 'starts_at', 'ends_at',
            'max_uses', 'used_count', 'min_order_amount', 'is_active',
            'product_ids', 'category_ids', 'created_at', 'updated_at',
        ]
        read_only_fields = ['id', 'used_count', 'created_at', 'updated_at']

    def validate_code(self, value):
        value = value.strip().upper()
        if not value:
            raise serializers.ValidationError('Mã giảm giá không được để trống.')
        queryset = DiscountCode.objects.filter(code=value)
        if self.instance:
            queryset = queryset.exclude(pk=self.instance.pk)
        if queryset.exists():
            raise serializers.ValidationError('Mã giảm giá đã tồn tại.')
        return value

    def validate(self, attrs):
        discount_type = attrs.get('discount_type', getattr(self.instance, 'discount_type', None))
        value = attrs.get('value', getattr(self.instance, 'value', None))
        starts = attrs.get('starts_at', getattr(self.instance, 'starts_at', None))
        ends = attrs.get('ends_at', getattr(self.instance, 'ends_at', None))
        if discount_type == DiscountCode.DiscountType.PERCENTAGE and (value is None or value < 0 or value > 100):
            raise serializers.ValidationError({'value': 'Phần trăm giảm phải từ 0 đến 100.'})
        if discount_type == DiscountCode.DiscountType.FIXED and (value is None or value <= 0):
            raise serializers.ValidationError({'value': 'Số tiền giảm phải lớn hơn 0.'})
        if starts and ends and starts >= ends:
            raise serializers.ValidationError({'ends_at': 'Thời gian kết thúc phải sau thời gian bắt đầu.'})
        if attrs.get('max_uses', getattr(self.instance, 'max_uses', None)) is not None and attrs.get('max_uses', getattr(self.instance, 'max_uses', None)) < 0:
            raise serializers.ValidationError({'max_uses': 'Giới hạn sử dụng không được âm.'})
        if attrs.get('min_order_amount', getattr(self.instance, 'min_order_amount', 0)) < 0:
            raise serializers.ValidationError({'min_order_amount': 'Giá trị đơn tối thiểu không được âm.'})
        return attrs


class AdminDiscountCodeViewSet(viewsets.ModelViewSet):
    queryset = DiscountCode.objects.all().prefetch_related('products', 'categories')
    serializer_class = DiscountCodeSerializer
    permission_classes = [OrderManagerPermission]
    http_method_names = ['get', 'post', 'patch', 'delete', 'head', 'options']


class AdminCustomerViewSet(viewsets.ModelViewSet):
    serializer_class = AdminCustomerSerializer
    permission_classes = [OrderManagerPermission]
    http_method_names = ['get', 'patch', 'head', 'options']

    def get_queryset(self):
        queryset = Customer.objects.annotate(
            purchase_count=Coalesce(Subquery(
                Order.objects.filter(
                    customer_phone=OuterRef('customer_phone'),
                ).exclude(status=Order.Status.CANCELLED).values(
                    'customer_phone',
                ).annotate(count=Count('id')).values('count'),
            ), 0),
            total_spent=Coalesce(Subquery(
                Order.objects.filter(
                    customer_phone=OuterRef('customer_phone'),
                ).exclude(status=Order.Status.CANCELLED).values(
                    'customer_phone',
                ).annotate(total=Sum('total_amount')).values('total'),
                output_field=DecimalField(max_digits=12, decimal_places=0),
            ), Decimal('0')),
        )
        search = self.request.query_params.get('search', '').strip()
        if search:
            queryset = queryset.filter(
                Q(customer_name__icontains=search)
                | Q(customer_phone__icontains=search)
            )
        return queryset

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        customer = self.get_object()
        serializer = self.get_serializer(
            customer, data=request.data, partial=partial,
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


def month_offset(year, month, offset):
    month_index = year * 12 + month - 1 + offset
    return divmod(month_index, 12)[0], divmod(month_index, 12)[1] + 1


def aware_start(value):
    return timezone.make_aware(
        datetime.combine(value, time.min),
        timezone.get_current_timezone(),
    )


class OrderCreateView(generics.CreateAPIView):
    serializer_class = OrderCreateSerializer
    permission_classes = [AllowAny]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        order = serializer.save()
        record_audit(
            request.user, AuditLog.Action.CREATED, 'order', order.id,
            f'Tạo đơn hàng #{order.id} từ cửa hàng.',
            {'status': {'before': None, 'after': order.status}},
        )
        response_serializer = OrderSerializer(
            order,
            context={'request': request},
        )

        return Response(
            response_serializer.data,
            status=status.HTTP_201_CREATED,
        )


class DiscountValidationView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = DiscountValidationSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        return Response({
            'discount_code': serializer.validated_data['discount_code'],
            'total_amount': serializer.validated_data['total_amount'],
            'discount_amount': serializer.validated_data['discount_amount'],
            'discounted_total': serializer.validated_data['discounted_total'],
        })


class OrderTrackingView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        order_id = request.query_params.get('order_id', '').strip().lstrip('#')
        phone = request.query_params.get('phone', '').strip()
        if not order_id.isdigit() or not phone:
            return Response(
                {'detail': 'Vui lòng cung cấp mã đơn và số điện thoại.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        order = Order.objects.filter(id=int(order_id), customer_phone=phone).prefetch_related('items').first()
        if not order:
            return Response(
                {'detail': 'Không tìm thấy đơn hàng với thông tin đã cung cấp.'},
                status=status.HTTP_404_NOT_FOUND,
            )
        return Response(OrderSerializer(order).data)


class CustomerOrderHistoryView(generics.ListAPIView):
    serializer_class = OrderSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Order.objects.filter(user=self.request.user).prefetch_related('items')


class DashboardStatsView(APIView):
    permission_classes = [OrderManagerPermission]

    def get(self, request):
        period = request.query_params.get('period', 'day')
        if period not in {'day', 'month'}:
            return Response(
                {'detail': 'period phải là day hoặc month.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if period == 'month':
            bucket_count = 12
            now = timezone.localdate()
            current_year, current_month = month_offset(
                now.year,
                now.month,
                0,
            )
            start_date = date(current_year, current_month, 1)
            previous_year, previous_month = month_offset(
                current_year,
                current_month,
                -bucket_count,
            )
            previous_start_date = date(previous_year, previous_month, 1)
            current_start = aware_start(start_date)
            previous_start = aware_start(previous_start_date)
            bucket_trunc = TruncMonth('created_at')
        else:
            bucket_count = 30
            end_date = timezone.localdate()
            start_date = end_date - timedelta(days=bucket_count - 1)
            previous_start_date = start_date - timedelta(days=bucket_count)
            current_start = aware_start(start_date)
            previous_start = aware_start(previous_start_date)
            bucket_trunc = TruncDate('created_at')

        sales_orders = Order.objects.exclude(status=Order.Status.CANCELLED)
        period_all_orders = Order.objects.filter(created_at__gte=current_start)
        period_orders = sales_orders.filter(created_at__gte=current_start)
        previous_orders = sales_orders.filter(
            created_at__gte=previous_start,
            created_at__lt=current_start,
        )

        period_revenue = period_orders.aggregate(
            total=Sum('total_amount'),
        )['total'] or Decimal('0')
        previous_revenue = previous_orders.aggregate(
            total=Sum('total_amount'),
        )['total'] or Decimal('0')
        period_order_count = period_orders.count()

        if previous_revenue:
            comparison_percent = round(
                (period_revenue - previous_revenue)
                * Decimal('100')
                / previous_revenue,
                2,
            )
        elif period_revenue:
            comparison_percent = Decimal('100')
        else:
            comparison_percent = Decimal('0')

        trend_rows = period_orders.annotate(
            bucket=bucket_trunc,
        ).values('bucket').annotate(
            revenue=Sum('total_amount'),
            orders=Count('id'),
        ).order_by('bucket')
        trend_map = {
            row['bucket'].strftime('%Y-%m' if period == 'month' else '%Y-%m-%d'): row
            for row in trend_rows
        }

        revenue_trend = []
        if period == 'month':
            for offset in range(bucket_count):
                year, month = month_offset(
                    start_date.year,
                    start_date.month,
                    offset,
                )
                label = f'{year:04d}-{month:02d}'
                row = trend_map.get(label, {})
                revenue_trend.append({
                    'label': label,
                    'revenue': row.get('revenue', Decimal('0')),
                    'orders': row.get('orders', 0),
                })
        else:
            for offset in range(bucket_count):
                current_date = start_date + timedelta(days=offset)
                label = current_date.isoformat()
                row = trend_map.get(label, {})
                revenue_trend.append({
                    'label': label,
                    'revenue': row.get('revenue', Decimal('0')),
                    'orders': row.get('orders', 0),
                })

        status_labels = dict(Order.Status.choices)
        status_rows = period_all_orders.values('status').annotate(
            count=Count('id'),
        )
        status_map = {row['status']: row['count'] for row in status_rows}
        orders_by_status = [
            {
                'status': value,
                'label': status_labels[value],
                'count': status_map.get(value, 0),
            }
            for value, _label in Order.Status.choices
        ]

        top_products = []
        item_rows = period_orders.filter(
            items__isnull=False,
        ).values(
            product_id=F('items__product_id'),
            product_name=F('items__product_name'),
        ).annotate(
            quantity_sold=Sum('items__quantity'),
            revenue=Sum(
                F('items__price') * F('items__quantity'),
                output_field=DecimalField(max_digits=12, decimal_places=0),
            ),
        ).order_by('-quantity_sold', '-revenue')[:5]
        for row in item_rows:
            top_products.append(row)

        low_stock_threshold = 5
        low_stock_products = list(
            Product.objects.filter(
                is_active=True,
                stock_quantity__lte=low_stock_threshold,
            ).select_related('category').order_by(
                'stock_quantity',
                'name',
            ).values(
                'id',
                'name',
                'stock_quantity',
                category_name=F('category__name'),
            )[:10]
        )

        total_orders = sales_orders.count()
        total_revenue = sales_orders.aggregate(
            total=Sum('total_amount'),
        )['total'] or Decimal('0')

        return Response({
            # Giữ các trường cũ để client cũ vẫn tương thích.
            'total_products': Product.objects.filter(is_active=True).count(),
            'total_orders': total_orders,
            'total_revenue': total_revenue,
            'period': period,
            'period_orders': period_order_count,
            'period_revenue': period_revenue,
            'average_order_value': (
                period_revenue / period_order_count
                if period_order_count else Decimal('0')
            ),
            'revenue_trend': revenue_trend,
            'orders_by_status': orders_by_status,
            'top_products': top_products,
            'low_stock_products': low_stock_products,
            'comparison': {
                'previous_revenue': previous_revenue,
                'change_percent': comparison_percent,
            },
        })


class ReportsView(APIView):
    permission_classes = [ReportPermission]

    def _date_range(self, request):
        end = request.query_params.get('end_date') or timezone.localdate().isoformat()
        start = request.query_params.get('start_date')
        if not start:
            start = (timezone.localdate() - timedelta(days=29)).isoformat()
        try:
            start_date = date.fromisoformat(start)
            end_date = date.fromisoformat(end)
        except ValueError:
            from rest_framework.exceptions import ValidationError
            raise ValidationError({'detail': 'Ngày bắt đầu/kết thúc không hợp lệ.'})
        if start_date > end_date:
            from rest_framework.exceptions import ValidationError
            raise ValidationError({'detail': 'Ngày bắt đầu phải trước ngày kết thúc.'})
        return start_date, end_date

    def _data(self, request):
        start_date, end_date = self._date_range(request)
        orders = Order.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date,
        ).exclude(status=Order.Status.CANCELLED)
        items = OrderItem.objects.filter(order__in=orders)
        revenue = orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
        cost = items.aggregate(
            total=Sum(
                F('cost_price') * F('quantity'),
                output_field=DecimalField(max_digits=14, decimal_places=0),
            ),
        )['total'] or Decimal('0')
        best_sellers = list(items.values(
            'product_id', 'product_name',
        ).annotate(
            quantity_sold=Sum('quantity'),
            revenue=Sum(
                F('price') * F('quantity'),
                output_field=DecimalField(max_digits=14, decimal_places=0),
            ),
            cost=Sum(
                F('cost_price') * F('quantity'),
                output_field=DecimalField(max_digits=14, decimal_places=0),
            ),
        ).order_by('-quantity_sold', '-revenue')[:20])
        for row in best_sellers:
            row['profit'] = row['revenue'] - row['cost']
        inventory = list(Product.objects.select_related('category').filter(
            is_active=True,
        ).values(
            'id', 'name', 'stock_quantity', 'cost_price', 'price',
            category_name=F('category__name'),
        ).order_by('stock_quantity', 'name'))
        for row in inventory:
            row['stock_value'] = row['stock_quantity'] * row['cost_price']
            row['retail_value'] = row['stock_quantity'] * row['price']
        return {
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat(),
            'summary': {
                'orders': orders.count(),
                'revenue': revenue,
                'cost': cost,
                'profit': revenue - cost,
            },
            'best_sellers': best_sellers,
            'inventory': inventory,
        }

    def get(self, request):
        report = self._data(request)
        export_format = request.query_params.get('export')
        if export_format in {'xlsx', 'pdf'}:
            return self._export(report, export_format)
        return Response(report)

    def _export(self, report, export_format):
        if export_format == 'xlsx':
            from openpyxl import Workbook
            workbook = Workbook()
            summary_sheet = workbook.active
            summary_sheet.title = 'Tổng quan'
            summary_sheet.append(['Báo cáo', 'Giá trị'])
            summary_sheet.append(['Từ ngày', report['start_date']])
            summary_sheet.append(['Đến ngày', report['end_date']])
            for label, key in [('Số đơn', 'orders'), ('Doanh thu', 'revenue'), ('Giá vốn', 'cost'), ('Lợi nhuận', 'profit')]:
                summary_sheet.append([label, report['summary'][key]])
            best_sheet = workbook.create_sheet('Sản phẩm bán chạy')
            best_sheet.append(['ID', 'Sản phẩm', 'Số lượng', 'Doanh thu', 'Giá vốn', 'Lợi nhuận'])
            for row in report['best_sellers']:
                best_sheet.append([row['product_id'], row['product_name'], row['quantity_sold'], row['revenue'], row['cost'], row['profit']])
            stock_sheet = workbook.create_sheet('Tồn kho')
            stock_sheet.append(['ID', 'Sản phẩm', 'Danh mục', 'Tồn kho', 'Giá vốn', 'Giá trị vốn', 'Giá bán', 'Giá trị bán lẻ'])
            for row in report['inventory']:
                stock_sheet.append([row['id'], row['name'], row['category_name'], row['stock_quantity'], row['cost_price'], row['stock_value'], row['price'], row['retail_value']])
            output = BytesIO()
            workbook.save(output)
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename="reports.xlsx"'
            return response

        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        output = BytesIO()
        document = SimpleDocTemplate(output, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        content = [
            Paragraph('Báo cáo kinh doanh', styles['Title']),
            Paragraph(f"{report['start_date']} đến {report['end_date']}", styles['Normal']),
            Spacer(1, 12),
        ]
        summary = report['summary']
        content.append(Table([
            ['Số đơn', 'Doanh thu', 'Giá vốn', 'Lợi nhuận'],
            [summary['orders'], f"{summary['revenue']:,}", f"{summary['cost']:,}", f"{summary['profit']:,}"],
        ], style=TableStyle([('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey)])))
        content.append(Spacer(1, 12))
        content.append(Paragraph('Sản phẩm bán chạy', styles['Heading2']))
        content.append(Table(
            [['Sản phẩm', 'Số lượng', 'Doanh thu', 'Lợi nhuận']] +
            [[row['product_name'], row['quantity_sold'], f"{row['revenue']:,}", f"{row['profit']:,}"] for row in report['best_sellers']],
            repeatRows=1,
            style=TableStyle([('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey)]),
        ))
        document.build(content)
        response = HttpResponse(output.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reports.pdf"'
        return response


class AdminOrderViewSet(viewsets.ModelViewSet):
    queryset = Order.objects.all().prefetch_related('items')
    serializer_class = AdminOrderSerializer
    permission_classes = [OrderManagerPermission]
    http_method_names = ['get', 'patch', 'post', 'head', 'options']

    @action(detail=False, methods=['get'], url_path='export', url_name='export')
    def export_orders(self, request):
        orders = self.get_queryset().order_by('id')
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['id', 'customer_name', 'customer_phone', 'shipping_address',
                         'status', 'payment_method', 'total_amount', 'items',
                         'created_at', 'updated_at'])
        for order in orders:
            items = '; '.join(
                f'{item.product_name} x {item.quantity}' for item in order.items.all()
            )
            writer.writerow([order.id, order.customer_name, order.customer_phone,
                             order.shipping_address, order.status, order.payment_method,
                             order.total_amount, items, order.created_at.isoformat(),
                             order.updated_at.isoformat()])
        response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="orders.csv"'
        return response

    @action(detail=False, methods=['get'], url_path='revenue-export', url_name='revenue-export')
    def export_revenue(self, request):
        queryset = self.get_queryset().exclude(status=Order.Status.CANCELLED)
        start = request.query_params.get('start_date')
        end = request.query_params.get('end_date')
        if start:
            queryset = queryset.filter(created_at__date__gte=start)
        if end:
            queryset = queryset.filter(created_at__date__lte=end)
        rows = queryset.annotate(day=TruncDate('created_at')).values('day').annotate(
            orders=Count('id'), revenue=Sum('total_amount')
        ).order_by('day')
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['date', 'orders', 'revenue'])
        for row in rows:
            writer.writerow([row['day'].isoformat(), row['orders'], row['revenue']])
        response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="revenue.csv"'
        return response

    def get_queryset(self):
        queryset = super().get_queryset().prefetch_related(
            'status_history__changed_by',
        )
        search = self.request.query_params.get('search', '').strip()
        if search:
            order_id_search = search.lstrip('#').strip()
            query = Q(
                customer_name__icontains=search,
            ) | Q(customer_phone__icontains=search)
            if order_id_search.isdigit():
                query |= Q(id=int(order_id_search))
            queryset = queryset.filter(query)
        if self.request.query_params.get('status'):
            queryset = queryset.filter(
                status=self.request.query_params['status'],
            )
        start = self.request.query_params.get(
            'start_date',
            self.request.query_params.get('date_from'),
        )
        end = self.request.query_params.get(
            'end_date',
            self.request.query_params.get('date_to'),
        )
        if start:
            queryset = queryset.filter(created_at__date__gte=start)
        if end:
            queryset = queryset.filter(created_at__date__lte=end)
        return queryset

    def _record_status_change(self, order, old_status, reason=''):
        if old_status != order.status:
            OrderStatusHistory.objects.create(
                order=order,
                from_status=old_status,
                to_status=order.status,
                reason=reason,
                changed_by=self.request.user,
            )

    def perform_update(self, serializer):
        old_status = serializer.instance.status
        order_instance = serializer.instance
        tracked = (
            'customer_name', 'customer_phone', 'shipping_address', 'status',
            'payment_method', 'internal_note', 'cancellation_reason',
        )
        before = {field: getattr(order_instance, field) for field in tracked}
        reason = serializer.validated_data.get('reason', '').strip()
        status_value = serializer.validated_data.get('status')
        save_kwargs = {}
        if status_value == Order.Status.CANCELLED:
            save_kwargs['cancellation_reason'] = reason
        order = serializer.save(**save_kwargs)
        self._record_status_change(order, old_status, reason)
        changes = {
            field: {'before': before[field], 'after': getattr(order, field)}
            for field in tracked if before[field] != getattr(order, field)
        }
        if reason and old_status == order.status:
            changes['reason'] = {'before': None, 'after': reason}
        if changes:
            record_audit(
                self.request.user, AuditLog.Action.UPDATED, 'order', order.id,
                f'Cập nhật đơn hàng #{order.id}.', changes,
            )

    @action(detail=True, methods=['post'])
    def confirm(self, request, pk=None):
        order = self.get_object()
        if order.status != Order.Status.PENDING:
            return Response(
                {'detail': 'Chỉ có thể xác nhận đơn đang chờ xác nhận.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        old_status = order.status
        reason = str(request.data.get('reason', '')).strip()
        order.status = Order.Status.CONFIRMED
        order.save(update_fields=['status', 'updated_at'])
        self._record_status_change(order, old_status, reason)
        record_audit(
            request.user, AuditLog.Action.UPDATED, 'order', order.id,
            f'Xác nhận đơn hàng #{order.id}.',
            {'status': {'before': old_status, 'after': order.status}},
        )
        return Response(self.get_serializer(order).data)

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        order = self.get_object()
        reason = str(request.data.get('reason', '')).strip()
        if not reason:
            return Response(
                {'reason': 'Vui lòng cung cấp lý do hủy đơn.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if order.status in {Order.Status.COMPLETED, Order.Status.CANCELLED}:
            return Response(
                {'detail': 'Không thể hủy đơn đã hoàn tất hoặc đã hủy.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        old_status = order.status
        order.status = Order.Status.CANCELLED
        order.cancellation_reason = reason
        order.save(update_fields=[
            'status',
            'cancellation_reason',
            'updated_at',
        ])
        self._record_status_change(order, old_status, reason)
        record_audit(
            request.user, AuditLog.Action.UPDATED, 'order', order.id,
            f'Hủy đơn hàng #{order.id}.',
            {
                'status': {'before': old_status, 'after': order.status},
                'cancellation_reason': {'before': '', 'after': reason},
            },
        )
        return Response(self.get_serializer(order).data)
